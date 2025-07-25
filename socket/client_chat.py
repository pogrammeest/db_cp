from utils import logger
from oop_socket import Socket
import asyncio
from datetime import datetime, date
from os import system
from exception import SocketException
from sys import platform
import time

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class Client(Socket):
    def __init__(self):
        self.chat_is_working = False
        self.messages = ""
        self.tasks = []
        self.is_sended = False
        super(Client, self).__init__()

    def set_up(self):
        try:
            self.socket.connect((self.address, self.port))
        except ConnectionError:
            for i in range(7, -1, -1):
                if platform == "win32":
                    system("cls")
                else:
                    system("clear")
                logger.info("Sorry, server is offline")
                logger.info("This program will close in %d seconds" % i)
                time.sleep(1)
            logger.info("Good bye...")
            time.sleep(1)
            exit(0)
        self.socket.setblocking(False)

    def registration(self):
        sending_data = {
            "name": "",
            "email": "",
            "password": "",
        }
        valid_pass = False

        input()

        logger.info("Начнем регистрацию!\nИмя:")
        sending_data["name"] = input()
        logger.info("\nПочта:")
        sending_data["email"] = input()
        while not valid_pass:
            logger.info("\nПароль:")
            sending_data["password"] = input()
            logger.info("\nповторите пароль:")
            repeat_pass = input()
            if sending_data["password"] == repeat_pass:
                valid_pass = True
            else:
                logger.info("Пароли не совпадают пожалуйста, повторите попытку!")
        logger.info("Вы успешно зарегистрированы!\n")
        logger.info("\n")
        logger.info(sending_data)
        input()
        return sending_data

    def create_xlsx(self, title, field_names, rows):
        # Создание книги и листа
        wb = Workbook()
        ws = wb.active
        ws.title = title

        # Запись заголовков
        ws.append(field_names)

        # Запись строк
        for row in rows:
            formatted_row = []
            for value in row:
                if isinstance(value, (datetime, date)):
                    formatted_row.append(
                        value.strftime("%Y-%m-%d %H:%M:%S")
                        if isinstance(value, datetime)
                        else value.strftime("%Y-%m-%d")
                    )
                else:
                    formatted_row.append(value)
            ws.append(formatted_row)

        for col_idx, column in enumerate(field_names, 1):
            max_length = max(
                len(str(ws.cell(row=row_idx, column=col_idx).value))
                for row_idx in range(1, len(rows) + 2)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        now = datetime.now()
        formatted = now.strftime("%Y-%m-%d_%H:%M:%S")
        name = f"./reports/{title}_{formatted}.xlsx"
        self.messages += f'Report with path: "{name}" was generated!\n'

        wb.save(name)

    async def listen_socket(self, listened_socket):
        while True:
            try:
                if not self.is_working:
                    return
                try:
                    if self.is_sended:
                        data = await asyncio.wait_for(
                            super(Client, self).listen_socket(listened_socket),
                            timeout=3,
                        )
                    else:
                        data = await super(Client, self).listen_socket(listened_socket)
                except asyncio.TimeoutError:
                    self.is_working = False
                    return
                except SocketException:
                    logger.info(
                        "Соединение с сервером потеряно, нажмите Enter, чтобы выйти из программы."
                    )
                    self.is_working = False
                    return
                if self.is_sended:
                    self.is_sended = False

                data = data.get("data", {})

                if data.get("command") == "disconnect":
                    return

                if data.get("root") == "server" and "request" in data:
                    if data["request"] == "chat":
                        if not self.chat_is_working:
                            self.chat_is_working = True
                            self.messages += f"$$SERVER MESSAGE$$: чат включен\n"
                        else:
                            self.chat_is_working = False
                            self.messages += f"$$SERVER MESSAGE$$: чат выключен\n"
                    elif data["request"] == "clear":
                        self.messages = "Экран очищен\n"
                    elif data["request"] == "show_db":
                        self.messages += data["message_text"] + "\n"
                    elif data["request"] == "reg":
                        self.messages = "Экран очищен\n"
                elif data.get("root") == "server":
                    self.messages += f"$$SERVER MESSAGE$$:{data['message_text']}\n"
                elif data.get("root") == "user" and self.chat_is_working:
                    self.messages += f"{data['message_time']}:{data['message_text']}\n"
                if data.get("root") == "server" and data.get("request") == "xlsx":
                    field_names = data.get("field_names")
                    rows = data.get("rows")
                    title = data.get("title")
                    self.create_xlsx(title, field_names, rows)

                if platform == "win32":
                    system("cls")
                else:
                    system("clear")
                logger.info(self.messages)
            except (SocketException, ConnectionError):
                logger.info("\n Server is offline...")
                time.sleep(5)
                exit(0)

    async def send_data(self, **kwargs):
        while True:
            if not self.is_working:
                return
            message = await self.main_loop.run_in_executor(None, input, "")
            encrypted_data = {
                "root": "user",
                "chat_is_working": self.chat_is_working,
                "message_text": message,
                "message_time": f"{datetime.now().hour}:{datetime.now().minute}:{datetime.now().second}",
            }
            if not self.chat_is_working:
                self.messages += "$$>" + message + "\n"
            await super(Client, self).send_data(where=self.socket, data=encrypted_data)
            self.is_sended = True
            if "/exit" == message:
                return

    async def shutdown(self):
        if not self.is_working:
            return
        logger.info("Завершаем работу клиента...")
        self.is_working = False
        for task in self.tasks:
            task.cancel()
        await asyncio.gather(*self.tasks, return_exceptions=True)
        try:
            self.socket.close()
        except Exception as e:
            logger.error(f"Ошибка при закрытии сокета: {e}")

    async def start_msg(self):
        data = {
            "root": "user",
            "chat_is_working": self.chat_is_working,
            "message_text": "/start",
            "message_time": f"{datetime.now().hour}:{datetime.now().minute}:{datetime.now().second}",
        }
        await super(Client, self).send_data(where=self.socket, data=data)

    async def main(self):
        await self.start_msg()
        listen_task = self.main_loop.create_task(self.listen_socket(self.socket))
        send_task = self.main_loop.create_task(self.send_data())
        self.tasks = [listen_task, send_task]

        try:
            await asyncio.gather(*self.tasks)
        except asyncio.CancelledError:
            logger.info("Main task cancelled")
        finally:
            await self.shutdown()


if __name__ == "__main__":
    client = Client()
    client.set_up()
    client.start()
